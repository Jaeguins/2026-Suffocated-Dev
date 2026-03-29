import sys
import os
import argparse
import openpyxl

HEADER_ROWS = 2  # row 0: type, row 1: name


def xlsx_to_dsv(input_path, output_dir, exclude_prefix, delimiter='\t', append=False):
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    sheets = [s for s in wb.sheetnames if not s.startswith(exclude_prefix)]

    if not sheets:
        print(f"[WARN] No sheets to export (all excluded by prefix '{exclude_prefix}').")
        wb.close()
        return

    os.makedirs(output_dir, exist_ok=True)

    for sheet_name in sheets:
        ws = wb[sheet_name]
        output_path = os.path.join(output_dir, sheet_name + '.tsv')

        rows = list(ws.iter_rows(values_only=True))

        if append and os.path.exists(output_path):
            data_rows = rows[HEADER_ROWS:]
            mode = 'a'
        else:
            data_rows = rows
            mode = 'w'

        with open(output_path, mode, encoding='utf-8', newline='') as f:
            for row in data_rows:
                line = delimiter.join('' if v is None else str(v) for v in row)
                f.write(line + '\n')

        action = 'appended' if mode == 'a' else 'created'
        print(f"[OK] '{sheet_name}' -> {output_path} ({action})")

    wb.close()


def main():
    parser = argparse.ArgumentParser(description='Convert Excel sheets to delimiter-separated files.')
    parser.add_argument('input',                help='Input directory containing .xlsx files')
    parser.add_argument('output',               help='Output directory')
    parser.add_argument('exclude_prefix',       help='Exclude sheets whose name starts with this text')
    parser.add_argument('-d', '--delimiter',    default='\t',
                        help='Delimiter character (default: tab)')
    args = parser.parse_args()

    if not os.path.isdir(args.input):
        print(f"[ERROR] Input path is not a directory: {args.input}")
        sys.exit(1)

    xlsx_files = [
        os.path.join(args.input, f)
        for f in os.listdir(args.input)
        if f.endswith('.xlsx')
    ]

    if not xlsx_files:
        print(f"[WARN] No .xlsx files found in '{args.input}'.")
        sys.exit(0)

    delimiter = args.delimiter.encode().decode('unicode_escape')

    # Clear output directory
    os.makedirs(args.output, exist_ok=True)
    for f in os.listdir(args.output):
        if not f.endswith('.tsv'):
            continue
        file_path = os.path.join(args.output, f)
        if os.path.isfile(file_path):
            os.remove(file_path)
    print(f"[OK] Cleared output directory: {args.output}")

    # Process all xlsx files in append mode
    for xlsx_path in xlsx_files:
        print(f"[--] Processing: {xlsx_path}")
        xlsx_to_dsv(xlsx_path, args.output, args.exclude_prefix, delimiter, append=True)


if __name__ == '__main__':
    main()
